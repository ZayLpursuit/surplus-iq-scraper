"""
Step 1: Import existing Excel → SQLite database
Run once to bootstrap your database from the Excel file already produced.

Usage: python3 import_excel.py
"""

import sqlite3
import openpyxl
import glob
import os
import sys
from datetime import datetime

DB_FILE = "surplus_leads.db"

def create_db(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS leads (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            case_number      TEXT UNIQUE,
            defendant        TEXT,
            plaintiff        TEXT,
            sheriff_number   TEXT,
            address          TEXT,
            sale_date        TEXT,
            approx_judgment  TEXT,
            status           TEXT,
            county           TEXT,
            detail_url       TEXT,
            first_seen       TEXT,
            last_updated     TEXT,
            is_new           INTEGER DEFAULT 1
        )
    """)
    conn.commit()
    print("  ✓ Database table ready")

def import_excel(conn, filepath):
    wb   = openpyxl.load_workbook(filepath)
    ws   = wb["All Sold Cases"]
    now  = datetime.now().isoformat()

    # Read headers from row 2 (row 1 is the title)
    headers = [cell.value for cell in ws[2]]
    print(f"  Columns found: {headers}")

    # Map Excel column names to DB columns
    col_map = {
        "Case #":          "case_number",
        "Defendant":       "defendant",
        "Plaintiff":       "plaintiff",
        "Sheriff #":       "sheriff_number",
        "Address":         "address",
        "Sale Date":       "sale_date",
        "Approx Judgment": "approx_judgment",
        "Status":          "status",
        "County":          "county",
        "Case Link":       "detail_url",
    }

    inserted = 0
    skipped  = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue

        record = {}
        for i, header in enumerate(headers):
            if header in col_map and i < len(row):
                db_col = col_map[header]
                val    = row[i]
                # Case Link cell shows "View →" but we need the actual URL
                if header == "Case Link" and val == "View →":
                    val = None
                record[db_col] = str(val).strip() if val else ""

        # Use case_number as unique key; fall back to sheriff_number
        key = record.get("case_number") or record.get("sheriff_number")
        if not key:
            skipped += 1
            continue

        record["case_number"]  = key
        record["first_seen"]   = now
        record["last_updated"] = now
        record["is_new"]       = 0  # existing data is not "new"

        try:
            conn.execute("""
                INSERT OR IGNORE INTO leads
                (case_number, defendant, plaintiff, sheriff_number, address,
                 sale_date, approx_judgment, status, county, detail_url,
                 first_seen, last_updated, is_new)
                VALUES
                (:case_number, :defendant, :plaintiff, :sheriff_number, :address,
                 :sale_date, :approx_judgment, :status, :county, :detail_url,
                 :first_seen, :last_updated, :is_new)
            """, record)
            if conn.execute("SELECT changes()").fetchone()[0]:
                inserted += 1
            else:
                skipped += 1
        except Exception as e:
            print(f"    ⚠  Row error: {e} — {record.get('case_number')}")
            skipped += 1

    conn.commit()
    return inserted, skipped

def main():
    print("=" * 50)
    print("  EXCEL → SQLITE IMPORTER")
    print("=" * 50)

    # Find the Excel file
    excel_files = sorted(glob.glob("nj_surplus_leads_*.xlsx"))
    if not excel_files:
        print("\n✗ No Excel file found in this folder.")
        print("  Make sure surplus_scraper.py ran successfully first.")
        sys.exit(1)

    excel_file = excel_files[-1]  # use most recent
    print(f"\n  Excel file: {excel_file}")
    print(f"  Database:   {DB_FILE}\n")

    conn = sqlite3.connect(DB_FILE)
    create_db(conn)

    print(f"  Importing rows...")
    inserted, skipped = import_excel(conn, excel_file)

    total = conn.execute("SELECT COUNT(*) FROM leads").fetchone()[0]
    conn.close()

    print(f"\n{'=' * 50}")
    print(f"  ✅  {inserted} rows imported")
    print(f"  ⏭   {skipped} duplicates skipped")
    print(f"  📦  {total} total leads in database")
    print(f"  📁  {DB_FILE} created in this folder")
    print(f"{'=' * 50}")

if __name__ == "__main__":
    main()
